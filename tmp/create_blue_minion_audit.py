from __future__ import annotations

import csv
from pathlib import Path
from typing import Final

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT: Final[Path] = Path(r"C:\Users\Focus on you\Desktop\projet app\dice_throne_survie")
XLSX_PATH: Final[Path] = ROOT / "audit_minions_bleus.xlsx"
CSV_PATH: Final[Path] = ROOT / "audit_minions_bleus.csv"

HEADERS: Final[list[str]] = [
    "Nom du minion",
    "Texte affichage ATK",
    "Interprétation du JSON pour l’attaque",
    "Analyse ATK (écart / mécanique non gérée)",
    "Texte affichage DEF",
    "Interprétation du JSON pour la défense",
    "Analyse IA DEF (affichage vs JSON)",
    "Tokens de l’ennemi",
    "Ce que l’app sait gérer pour ces tokens",
]

ROWS: Final[list[list[str]]] = [
    [
        "Maraud Bestial (bleu-bleu-001)",
        "Résumé combat : 2B + 2O + 1R → 6 dégâts imparables. Détail : Sauvagerie de l’Âme.",
        "Objectif 2 blancs + 2 orange + 1 rouge. Effet : 6 dégâts, undefendable=true.",
        "Conforme. Dégâts et caractère imparable sont codés explicitement et appliqués.",
        "Résumé combat : chaque O prévient 1 dégât, répétable.",
        "4 dés. Pour chaque orange : prevent=1, repeat=true.",
        "Conforme. Affichage et résolution automatique correspondent au defensePlan.",
        "Aucun token.",
        "Sans objet.",
    ],
    [
        "Archer Aveugle (bleu-bleu-002)",
        "Résumé combat : 3O→4, 4O→5, 5O→6, tous imparables. Rappel conditionnel : 3 valeurs identiques → Eboulissement.",
        "Trois paliers orange donnant 4/5/6 dégâts imparables. Si l’attaque réussit et contient 3 valeurs identiques : inflige Eboulissement (règle surtout textuelle dans le JSON).",
        "Conforme grâce à du code spécifique. La règle Eboulissement n’est pas correctement structurée dans conditionalRules.tokens, mais le code la compense.",
        "Résumé combat : 2O préviennent 4 dégâts.",
        "4 dés. Au moins 2 orange : prévient exactement 4, non répétable.",
        "Conforme et automatisé.",
        "Eboulissement (alias de Blinding Light / Lumière aveuglante).",
        "Le catalogue reconnaît l’alias et l’affiche. L’attaque de cet ennemi sait l’ajouter. La mécanique complète du token reste partielle et dépend de code spécifique.",
    ],
    [
        "Mage de l’Entropie (bleu-bleu-003)",
        "Résumé combat : objectif 2O + 1R, puis phase de dé supplémentaire. Détail : 7 + Chaos ; blanc gagne 2 Chaos ; orange inflige Sort 6.",
        "Objectif 2O+1R. Extra-roll : blanc → 2 Chaos au minion ; orange → Sort 6 au héros ; dégâts finaux = 7 + nombre de Chaos.",
        "Conforme grâce à un résolveur dédié. Le moteur générique ne sait pas interpréter cette formule ni l’extra-roll.",
        "Résumé combat : chaque O donne Chaos puis inflige 1 dégât imparable par Chaos.",
        "4 dés. Par orange : gagne Chaos ; dégâts de retour imparables selon le total de Chaos ; répétable.",
        "Conforme grâce à du code spécifique.",
        "Chaos ; Sort 6.",
        "Chaos est géré spécifiquement pour cet ennemi ; Sort 6 peut être ajouté au héros mais sa mécanique générale n’est pas exécutée. Le catalogue indique maxStack=1 pour Chaos, incohérent avec le gain de 2 et les mécaniques à 3 piles.",
    ],
    [
        "Mage de Sang (bleu-bleu-004)",
        "Résumé combat standard : 3O/4O/5O affichent 0 dégât ; le détail indique vol de 3/4/5 PV.",
        "3O → vole 3 PV ; 4O → vole 4 PV ; 5O → vole 5 PV. damage=0, stealHp=3/4/5.",
        "Résolution conforme grâce à du code spécifique, mais le résumé ATK est trompeur car les badges de dégâts montrent 0 au lieu du vol de vie.",
        "Résumé combat : chaque O vole 1 PV ; au moins 1R donne Chaos.",
        "3 dés. Chaque orange vole 1 PV (répétable). Un rouge donne Chaos (non répétable).",
        "Conforme. Plusieurs rouges ne donnent qu’un Chaos, comme prévu par le code.",
        "Chaos.",
        "Gestion spécifique complète : gain en défense et à l’entretien ; à 3 Chaos, dépense 3 et vole 3 PV. Le maxStack=1 du catalogue est incohérent et non imposé par les ajouts programmatiques.",
    ],
    [
        "Sorcière d’Os (bleu-bleu-005)",
        "Résumé combat : 1B + 2O + 1R → 3 dégâts imparables ; rappels Silence/Parasite. Détail : Dépérissement + Silence + Parasite + 3 imparables ; entretien Siphon vital.",
        "Action principale : 3 dégâts imparables, tokens Silence et Parasite. Dépérissement n’est pas dans actions.tokens. Passif textuel : gagne Siphon vital.",
        "Seuls les 3 dégâts imparables sont appliqués. Silence et Parasite sont parsés mais non propagés ; Dépérissement et Siphon vital ne sont pas exécutés.",
        "Résumé combat : vide. Fiche détail : « Defense: » sans règle.",
        "4 dés. Par orange : vole 1 PV. Par rouge : inflige Hémorragie. Le JSON encode mal le vol (stealHp=0) et utilise des conditions textuelles.",
        "Écart critique : defensePlan non parsé ; aucune défense automatique ni affichage utile.",
        "Dépérissement ; Silence ; Parasite ; Siphon vital ; Hémorragie.",
        "Catalogue/UI : tous reconnus. Dépérissement et Hémorragie sont marqués supportés, mais cet ennemi ne les applique pas. Silence, Parasite et Siphon vital sont non supportés. Hémorragie française ne déclenche pas le handler exact « Bleed ».",
    ],
    [
        "Cyclope Brutal (bleu-bleu-006)",
        "Résumé combat : 4B + 1R → 6 dégâts. Détail : 6 + valeur d’un dé ; en cas d’échec, 2 dégâts imparables.",
        "Action structurée : 6 dégâts. Le jet supplémentaire et l’ajout de sa valeur sont seulement textuels ; passif d’échec : 2 dégâts imparables.",
        "Seuls 6 dégâts défendables sont appliqués. Extra-roll, ajout de valeur et effet d’échec non gérés.",
        "Résumé combat : vide.",
        "1 dé. Blanc → retourne 3 dégâts ; orange → soigne 2 PV.",
        "Non géré : defensePlan ignoré, aucun effet automatique.",
        "Aucun token.",
        "Sans objet.",
    ],
    [
        "The Hermit (bleu-bleu-007)",
        "Résumé combat : 1B + 3O → 3 dégâts imparables ; rappel Dépérissement/Brûlure. Détail : dé supplémentaire, blanc +2 dégâts, orange Dépérissement, rouge Brûlure.",
        "Base : 3 dégâts imparables. Issues déclarées : blanc +2 dégâts ; orange Dépérissement ; rouge Brûlure, mais sans relation extraRoll exécutable.",
        "Seule la base de 3 dégâts imparables est appliquée. Dé supplémentaire et trois issues ignorés.",
        "Résumé combat : vide.",
        "5 dés. Condition numérique 2 → prévient 3 dégâts.",
        "Non géré : condition number et defensePlan ignorés.",
        "Dépérissement ; Brûlure.",
        "Les deux sont reconnus par le catalogue. Aucun n’est infligé ici. Le moteur d’entretien attend « Burn », donc une chaîne « Brûlure » resterait inactive ; Dépérissement n’a pas de réduction automatique fiable dans ce chemin.",
    ],
    [
        "Dark Specter (bleu-bleu-008)",
        "Résumé combat : micro-suite → 4 ; petite suite → 0 ; grande suite → 0. Rappels Enchevêtrement/Silence. Détail : micro 4+Enchevêtrement, petite 6+Silence, grande 8+Sort 6.",
        "Micro (longueur 3) : 4 + Enchevêtrement. Petite (4) : Silence mais damage=0. Grande (5) : damage=0 et aucun token.",
        "Micro : seuls 4 dégâts. Petite/grande : zéro car les effets structurés à 0 bloquent le fallback texte. Tous les tokens sont ignorés.",
        "Résumé combat : vide.",
        "3 dés. Texte : prévient 2 dégâts par orange ; structure prevent=2 mais répétition mal décrite.",
        "Non géré : défense vide et aucun calcul automatique.",
        "Enchevêtrement ; Silence ; Sort 6.",
        "Catalogue/UI : reconnus. Enchevêtrement est marqué supporté, mais aucune réduction générique du nombre de jets ; Silence et Sort 6 sont non supportés. Aucun n’est appliqué par cette attaque.",
    ],
    [
        "The Butcher (bleu-bleu-009)",
        "Résumé combat : 1B + 2O + 1R → 3 dégâts imparables. Détail : dé supplémentaire ; orange → Commotion ; échec → soigne 2.",
        "Action principale : 3 dégâts imparables. Issue orange et passif de soin existent surtout en texte/structure dissociée ; aucun extraRoll relié.",
        "Seuls 3 dégâts imparables sont appliqués. Dé, Commotion et soin sur échec ignorés.",
        "Résumé combat : vide.",
        "1 dé. Orange → retourne 3 dégâts.",
        "Non géré : aucun retour de dégâts automatique.",
        "Commotion.",
        "Token reconnu mais appSupported=false. Quelques rappels UI existent, mais cet ennemi ne l’inflige pas et le cycle complet (saut de revenu puis retrait) n’est pas automatisé.",
    ],
    [
        "Vipère Vicieuse (bleu-bleu-010)",
        "Résumé combat : objectifs 3B et 3B+1R ; 7 dégâts ; rappel Poison.",
        "Texte : 3B→7 ; 3B+R→Poison+7. JSON : une seule action 3B→7 contenant déjà Poison ; aucune action distincte pour le palier avec rouge.",
        "7 dégâts sont appliqués. Poison n’est pas propagé. Données incohérentes : Poison est rattaché au mauvais objectif.",
        "Résumé combat : vide.",
        "4 dés. Un rouge inflige Poison.",
        "Non géré : le defensePlan n’est pas interprété.",
        "Poison.",
        "Le moteur sait traiter exactement « Poison » à l’entretien (-1 PV par pile), et le catalogue limite à 3. Mais cet ennemi ne l’ajoute automatiquement ni en attaque ni en défense.",
    ],
    [
        "Vaurien (bleu-bleu-011)",
        "Résumé combat : petite suite → 0 ; grande suite → 0. Détail : moitié des CP du minion en dégâts / tous ses CP en dégâts ; début de tour +2 CP.",
        "Petite suite : formule ceil(minionCp/2), damage=0. Grande : effet nul. Action textuelle de début de tour : +2 CP, non structurée.",
        "Formules ignorées : deux suites à 0. Le +2 CP n’est pas exécuté ; seul le +1 CP générique d’entretien peut s’appliquer.",
        "Résumé combat : vide.",
        "4 dés. Condition number=2 : vole 1 CP ; deuxième condition identique sans effet.",
        "Non géré : aucune interprétation de la condition ou du vol de CP.",
        "Première Frappe (initial).",
        "Traitement ad hoc : l’ennemi commence en premier. Le token est appSupported=false et sa consommation n’est pas généralisée ; la comparaison dépend du libellé exact.",
    ],
    [
        "Épée Fantôme (bleu-bleu-012)",
        "Résumé combat : 3B/4B/5B → 5/6/7 dégâts (paliers bas via fallback) ; rappel Parasite.",
        "Trois objectifs 3B, 4B, 5B. Seule l’action 5B→7 est structurée. Règle conditionnelle : 3 valeurs identiques → Parasite.",
        "Dégâts probablement récupérés par parsing textuel pour 3B/4B, donc fragile. 5B→7 fiable. Parasite non appliqué.",
        "Résumé combat : vide.",
        "1 dé quelconque. Retourne ceil(valeur du dé / 2) dégâts via formule.",
        "Non géré : formule et condition any ignorées.",
        "Parasite.",
        "Reconnu pour l’affichage, appSupported=false. Aucune mécanique ni application automatique pour cet ennemi.",
    ],
    [
        "Elfe Flétri (bleu-bleu-013)",
        "Résumé combat : micro/petite/grande suite → 0/0/0 ; rappel Parasite. Détail textuel : 5 / Parasite+7 / Dépérissement+9.",
        "Les trois actions ont damage=0. Petite contient Parasite ; micro n’a aucun token ; grande omet Dépérissement.",
        "Attaque largement incorrecte : le zéro structuré empêche le fallback texte. Aucun dégât ni token n’est appliqué.",
        "Résumé combat : vide.",
        "5 dés. Condition number=2, tous les effets à zéro.",
        "Aucune mécanique exploitable dans les données et aucun traitement automatique.",
        "Parasite ; Dépérissement.",
        "Tous deux reconnus visuellement. Parasite est non supporté ; Dépérissement est marqué supporté mais sans effet automatique fiable ici. Aucun n’est appliqué.",
    ],
    [
        "Banshie Hurlante (bleu-bleu-014)",
        "Résumé combat : objectifs 2R/3R/4R/5R, valeurs 2/3/4/5 selon fallback ; rappel Silence. Détail : dégâts collatéraux à tous les adversaires, Silence si un seul adversaire.",
        "Quatre objectifs rouges. Seule l’action 5R→5 est structurée. conditionalRule Silence si un seul adversaire.",
        "Valeurs basses dépendantes du parsing texte. Les dégâts sont appliqués au héros actif seulement, pas à tous. Silence conditionnel ignoré.",
        "Résumé combat : vide.",
        "2 dés. Condition 1 rouge, effets à zéro.",
        "Aucun effet défensif exploitable ou automatisé.",
        "Silence.",
        "Reconnu pour l’UI, appSupported=false. Quelques contrôles de Silence existent ailleurs, mais cet ennemi ne l’inflige pas automatiquement.",
    ],
    [
        "Centaure Enragé (bleu-bleu-015)",
        "Résumé combat : 2B+1O→4 ; 2B+2O→5 ; 2B+2O+1R→6 attendu. Détail : défausse un token positif aléatoire.",
        "Trois objectifs. Seule l’action médiane 2B+2O→5 est structurée ; action textuelle de défausse sans effet.",
        "Palier 4 via fallback, palier 5 fiable, palier 6 incertain/non résolu. Suppression aléatoire d’un token positif non gérée.",
        "Résumé combat : vide.",
        "3 dés. Un orange inflige À Terre.",
        "Non géré : À Terre n’est pas appliqué par la défense.",
        "Première Frappe (initial) ; À Terre.",
        "Première Frappe modifie l’ordre initial. À Terre est reconnu et marqué supporté dans le catalogue, mais aucune mécanique complète ni application automatique depuis cette défense.",
    ],
    [
        "Chevalier à la Hache (bleu-bleu-016)",
        "Résumé combat : petite suite→6 ; grande suite→9. Détail : soigne 2 ; en cas d’échec gagne 3 Dégâts Bonus.",
        "Petite/grande suite : damage=6/9 et heal=2. Action d’échec textuelle encodée de façon incohérente avec damage=3 + token « Dégâts Bonus ».",
        "Dégâts 6/9 appliqués. Soin parsé mais ignoré. Échec ignoré. Token générique introuvable dans le catalogue.",
        "Résumé combat : vide.",
        "4 dés. Blanc retourne 1 ; orange prévient 2 ; rouge prévient 1, mais structure partielle.",
        "Non géré : aucun effet de défense automatique.",
        "Dégâts Bonus.",
        "Aucune entrée exacte. Le catalogue ne contient que Dégat Bonus 1 à 4, tous appSupported=false. Affichage en fallback possible, aucune sémantique.",
    ],
    [
        "Panthère Ténébreuse (bleu-bleu-017)",
        "Résumé combat : 3B/4B/5B→6/7/8 ; rappel Hémorragie.",
        "Trois objectifs ; seule action 5B→8 structurée. Trois valeurs identiques → Hémorragie via conditionalRule.",
        "Dégâts 6/7 via fallback et 8 fiable. Hémorragie conditionnelle ignorée.",
        "Résumé combat : vide.",
        "1 dé. Blanc → Hémorragie ; rouge → 2 Hémorragies, mais le JSON encode returnDamage=2 + une seule Hémorragie.",
        "Non géré et données incohérentes : aucun token ni retour automatique.",
        "Première Frappe (initial) ; Hémorragie.",
        "Première Frappe gère l’ordre. Hémorragie est reconnue visuellement mais non appliquée ; même ajoutée en français, elle ne déclenche pas le handler exact « Bleed ».",
    ],
    [
        "Harpie Cornue (bleu-bleu-018)",
        "Résumé combat : 2O/3O/4O/5O→3/4/5/6 dégâts imparables.",
        "Quatre objectifs orange ; seule action 5O→6 imparable structurée. Les autres valeurs sont dans le texte.",
        "Dégâts et caractère imparable sont récupérés, mais les paliers bas dépendent du fallback textuel.",
        "Résumé combat : vide.",
        "3 dés. Conditions 1O et number=2, effets tous à zéro.",
        "Aucun effet défensif déclaré de manière exploitable ou automatisé.",
        "Première Frappe (initial).",
        "Gère l’ordre initial uniquement ; token non supporté génériquement et consommation non généralisée.",
    ],
    [
        "Elfe Agile (bleu-bleu-019)",
        "Résumé combat : 3B + 2O → 6 dégâts. Détail : Eboulissement +6 ; début de tour retire toutes les altérations positives du joueur.",
        "Action principale : 6 dégâts sans token. Action textuelle de purge des tokens positifs avec effets nuls.",
        "Seuls 6 dégâts sont appliqués. Eboulissement et suppression des tokens positifs ignorés.",
        "Résumé combat : vide.",
        "1 dé. Soigne selon la valeur du dé et le nombre de minions en jeu ; heal=0, aucune formule structurée.",
        "Non géré : formule absente et defensePlan ignoré.",
        "Eboulissement.",
        "Alias reconnu et appSupported=true, mais cet ennemi ne l’ajoute pas. Le comportement complet du token n’est pas piloté par le catalogue.",
    ],
    [
        "Farceur (bleu-bleu-020)",
        "Résumé combat : 1B + 2O + 1R → 4 dégâts imparables. Détail : dé supplémentaire ; blanc vole 2 CP ; orange défausse 1 carte ; rouge fait les deux.",
        "Action principale : 4 imparables. Actions séparées : blanc stealCp=2 ; orange discardCards=1 ; rouge effets nuls. Aucun extraRoll relié.",
        "Seuls 4 dégâts imparables. Dé supplémentaire, vol de CP et défausse ignorés.",
        "Résumé combat : vide.",
        "5 dés. Condition number=2 → prévient 3 dégâts.",
        "Non géré : condition number et prévention ignorées.",
        "Première Frappe (initial).",
        "Gère l’ordre initial. Pas de consommation générique ni de normalisation de libellé.",
    ],
    [
        "Bandit sans Âme (bleu-bleu-021)",
        "Résumé combat : 3B/4B/5B→5/6/7. Détail : 3 valeurs identiques → tous les adversaires engagés perdent 1 CP.",
        "Trois objectifs ; seule action 5B→7 structurée. Règle conditionnelle de perte de CP stockée en texte avec effets nuls.",
        "Dégâts récupérés, paliers bas via fallback. Perte de CP et portée multi-adversaires ignorées.",
        "Résumé combat : vide.",
        "3 dés. Retourne 1 dégât par blanc et par rouge ; structure réduite à returnDamage=1 sans répétition/formule.",
        "Non géré : aucun calcul de retour automatique.",
        "Première Frappe (initial).",
        "Gère seulement l’ordre initial.",
    ],
    [
        "Vibra l’Ésotérique (bleu-bleu-022)",
        "Résumé combat : 2B + 1/2/3O → 4/5/6 dégâts. Détail : début de tour gagne Chaos ; à 3 Chaos, les dépense et se soigne du montant des dégâts.",
        "Trois objectifs ; seule action maximale 2B+3O→6 structurée. Action de début de tour textuelle et vide.",
        "Dégâts 4/5 via fallback, 6 fiable. Cycle Chaos et soin dynamique entièrement ignorés.",
        "Résumé combat : vide.",
        "4 dés. Vole 1 PV par orange, mais stealHp=0 dans la structure.",
        "Non géré : vol absent et defensePlan ignoré.",
        "Première Frappe (initial) ; Chaos.",
        "Première Frappe gère l’ordre. Chaos n’a aucun traitement spécifique pour Vibra, contrairement aux profils 003/004 ; maxStack catalogue incohérent avec le seuil de 3.",
    ],
    [
        "Yokai (bleu-bleu-023)",
        "Résumé combat : 4O → 0. Détail : lancer un dé ; blanc 4 imparables ; orange vole 3 PV ; rouge vole 4 PV ; échec Silence + Sort 6.",
        "Action principale 4O : damage=0, extraRoll=null. Résultats B/O/R stockés comme actions sœurs. Action d’échec contient seulement Silence ; Sort 6 absent.",
        "Attaque validée mais effet nul. Aucun dé supplémentaire, aucun dégât/vol, aucun token d’échec.",
        "Résumé combat : vide.",
        "3 dés. Vole 1 PV par orange, mais stealHp=0.",
        "Non géré : aucun vol automatique.",
        "Silence ; Sort 6.",
        "Tous deux reconnus mais appSupported=false. Aucun n’est appliqué ; Sort 6 est en plus absent de l’action structurée d’échec.",
    ],
    [
        "Level 2 Minion (blue-generic)",
        "Résumé combat : texte brut « Precise strike: 4 damage » et « Pressure: -1 CP » ; pas de résolution automatique fiable.",
        "Plan style=none. Action textuelle à 4 dégâts ; Pressure a stealCp=0.",
        "Profil de repli, pas une carte bleue numérotée. Le style none retourne sans résolution : attaque manuelle/affichée seulement.",
        "Résumé combat : vide.",
        "2 dés. Texte « Blocks 2 damage », mais prevent=0.",
        "Non automatisé et probablement vide, car le champ top-level defense n’est pas alimenté.",
        "Aucun token.",
        "Sans objet.",
    ],
]


def status_fill(text: str) -> PatternFill:
    """Retourne une couleur d'analyse selon le niveau de prise en charge."""
    normalized = text.lower()
    if any(word in normalized for word in ("critique", "entièrement ignor", "effet nul", "largement incorrect", "non géré")):
        return PatternFill("solid", fgColor="F4CCCC")
    if any(word in normalized for word in ("partiel", "fragile", "fallback", "trompeur", "incohérent", "ignoré", "seuls")):
        return PatternFill("solid", fgColor="FCE5CD")
    if any(word in normalized for word in ("conforme", "automatisé", "fiable")):
        return PatternFill("solid", fgColor="D9EAD3")
    return PatternFill("solid", fgColor="FFF2CC")


def build_workbook() -> None:
    """Crée le classeur d'audit professionnel et le fichier CSV équivalent."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Audit minions bleus"

    sheet.merge_cells("A1:I1")
    sheet["A1"] = "Audit ATK / DEF / tokens — minions bleus"
    sheet["A1"].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28

    sheet.merge_cells("A2:I2")
    sheet["A2"] = (
        "Périmètre : 23 cartes bleu-bleu-001 à 023 + profil de repli blue-generic. "
        "Comparaison entre l’affichage réel de l’application, docs/enemy_profiles.json et le comportement Dart implémenté."
    )
    sheet["A2"].font = Font(name="Arial", size=10, italic=True, color="404040")
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 34

    sheet.merge_cells("A3:I3")
    sheet["A3"] = (
        "Sources : docs/enemy_profiles.json ; assets/data/token_catalog.json ; "
        "lib/data/enemy_profile_repository.dart ; lib/models/enemy_profile.dart ; "
        "lib/parts/fight.dart ; lib/game_engine.dart ; lib/main.dart. Audit réalisé sur la version 1.3.50+94."
    )
    sheet["A3"].font = Font(name="Arial", size=9, color="666666")
    sheet["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[3].height = 32

    header_row = 5
    for column, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=header_row, column=column, value=header)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[header_row].height = 45

    thin_gray = Side(style="thin", color="B7B7B7")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for row_index, row in enumerate(ROWS, start=header_row + 1):
        for column_index, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if column_index in (4, 7, 9):
                cell.fill = status_fill(value)
            elif row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F7F9FC")
        sheet.row_dimensions[row_index].height = 105

    widths = [31, 48, 52, 50, 42, 50, 48, 31, 54]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "B6"
    sheet.auto_filter.ref = f"A{header_row}:I{header_row + len(ROWS)}"
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = f"1:{header_row}"
    sheet.auto_filter.ref = f"A{header_row}:I{header_row + len(ROWS)}"

    table_ref = f"A{header_row}:I{header_row + len(ROWS)}"
    table = Table(displayName="AuditMinionsBleus", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    sheet["A5"].comment = Comment(
        "B = symbole blanc ; O = symbole orange/jaune ; R = symbole rouge.",
        "Claude",
    )

    legend = workbook.create_sheet("Légende et constats")
    legend.sheet_view.showGridLines = False
    legend.column_dimensions["A"].width = 27
    legend.column_dimensions["B"].width = 115
    legend["A1"] = "Légende / constat"
    legend["B1"] = "Explication"
    for cell in legend[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    legend_rows = [
        ("Vert", "Comportement affiché, données JSON et résolution automatique globalement conformes."),
        ("Orange", "Prise en charge partielle, fallback textuel, code spécifique ou incohérence limitée."),
        ("Rouge", "Mécanique absente, effet nul, défense non gérée ou écart critique."),
        ("B / O / R", "B = blanc ; O = orange (parfois nommé jaune dans les textes) ; R = rouge."),
        ("Affichage ATK", "Le résumé de combat reconstruit les objectifs et dégâts depuis attackPlan ; la fiche détail montre aussi les chaînes attacks brutes."),
        ("Affichage DEF", "Pour les profils 005 à 023, defensePlan n’est pas chargé dans EnemyProfile : le résumé est généralement vide."),
        ("Défenses automatisées", "Parmi les cartes bleues, seules 001 à 004 disposent de branches de résolution et d’affichage spécifiques."),
        ("Tokens", "Le catalogue sait reconnaître des alias pour l’UI, mais l’état stocke des chaînes brutes et l’entretien compare souvent des noms anglais exacts."),
        ("appSupported", "Indicateur informatif du catalogue : il ne garantit ni l’application du token, ni son effet, ni sa consommation."),
        ("Profil générique", "blue-generic est un profil de repli de rang bleu, pas une carte bleu-bleu numérotée."),
    ]
    for row_index, (label, explanation) in enumerate(legend_rows, start=2):
        legend.cell(row=row_index, column=1, value=label)
        legend.cell(row=row_index, column=2, value=explanation)
        for cell in legend[row_index]:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        legend.row_dimensions[row_index].height = 38

    legend["A2"].fill = PatternFill("solid", fgColor="D9EAD3")
    legend["A3"].fill = PatternFill("solid", fgColor="FCE5CD")
    legend["A4"].fill = PatternFill("solid", fgColor="F4CCCC")
    legend.freeze_panes = "A2"

    workbook.save(XLSX_PATH)

    with CSV_PATH.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle, delimiter=";")
        writer.writerow(HEADERS)
        writer.writerows(ROWS)


def validate_workbook() -> None:
    """Valide les dimensions et le contenu essentiel du classeur généré."""
    workbook = load_workbook(XLSX_PATH, data_only=False)
    sheet = workbook["Audit minions bleus"]
    assert sheet.max_row == 5 + len(ROWS)
    assert sheet.max_column == len(HEADERS)
    assert sheet["A6"].value == "Maraud Bestial (bleu-bleu-001)"
    assert sheet[f"A{5 + len(ROWS)}"].value == "Level 2 Minion (blue-generic)"
    assert "Légende et constats" in workbook.sheetnames
    assert XLSX_PATH.stat().st_size > 10_000
    assert CSV_PATH.stat().st_size > 10_000


if __name__ == "__main__":
    build_workbook()
    validate_workbook()
    print(f"Créé : {XLSX_PATH}")
    print(f"Créé : {CSV_PATH}")
    print(f"Lignes de données : {len(ROWS)}")
